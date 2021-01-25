# -*- coding:utf-8 -*-

"""
      ┏┛ ┻━━━━━┛ ┻┓
      ┃　　　　　　 ┃
      ┃　　　━　　　┃
      ┃　┳┛　  ┗┳　┃
      ┃　　　　　　 ┃
      ┃　　　┻　　　┃
      ┃　　　　　　 ┃
      ┗━┓　　　┏━━━┛
        ┃　　　┃   神兽保佑
        ┃　　　┃   代码无BUG！
        ┃　　　┗━━━━━━━━━┓
        ┃　　　　　　　    ┣┓
        ┃　　　　         ┏┛
        ┗━┓ ┓ ┏━━━┳ ┓ ┏━┛
          ┃ ┫ ┫   ┃ ┫ ┫
          ┗━┻━┛   ┗━┻━┛
"""

import openpyxl
from tqdm import tqdm
letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
          'W', 'X', 'Y', 'Z', 'AA']


def write_excel(file_name, data_list, col_name_list=[]):
    """

    :param file_name: 文件名
    :param data_list: 需要保存的数据（列表）
    :param col_name_list: 表头（如果确定该文件已经存在可以不传）
    :return:
    """
    try:
        wb = openpyxl.load_workbook(file_name)
        sheet1 = wb['Sheet1']
        for each in tqdm(data_list):
            col_len = len(sheet1['A'])
            # sheet1['A' + str(col_len + 1)] = str(col_len)
            for i in range(len(each)):
                tmp_str = letter[i] + str(col_len + 1)
                sheet1[tmp_str] = each[i]
        wb.save(file_name)
        return col_len
    
    except:
        wb = openpyxl.Workbook()
        wb.create_sheet('Sheet1', 0)
        sheet1 = wb['Sheet1']
        for i in range(len(col_name_list)):
            tmp_str = letter[i] + str(1)
            sheet1[tmp_str] = col_name_list[i]
        wb.save(file_name)
        return write_excel(file_name, data_list)


if __name__ == '__main__':
    pass
